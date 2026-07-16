# Lê a exportação do "Analisador de Trabalho > Colheita" do OPC (uma linha por
# máquina/operador/talhão) e gera o JSON da FROTA usado pelos cards do painel.
#
# Uso:  python parse_opc_frota.py data/opc_frota_milho_2026.xlsx data/opc_frota_milho_2026.json
#
# Agregação por máquina: soma área, tempo e combustível e SÓ ENTÃO divide —
# média de médias daria número errado. Operador da máquina = o que mais colheu (ha).
import json
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl


def norm(s):
    """minúsculas, sem acento, espaços colapsados — p/ casar cabeçalhos"""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip().lower()


# cabeçalho (normalizado) -> chave interna
COLS = {
    'talhoes': 'talhao',
    'nome da maquina': 'apelido',
    'equipamento': 'vin',
    'operadores': 'operador',
    'tipo de cultura': 'cultura',
    'area colhida': 'area',
    'umidade': 'umidade',
    'tempo de colheita': 'tempo',      # segundos
    'velocidade': 'velocidade',        # km/h
    'combustivel total': 'combustivel',  # litros
}


def ca_from_apelido(ap):
    """'CA S790 10' -> 'CA10' ; 'CA S770 21' -> 'CA21'. Sem número, devolve o apelido."""
    if not ap:
        return None
    m = re.search(r'(\d+)\s*$', str(ap).strip())
    return 'CA%02d' % int(m.group(1)) if m else str(ap).strip()


# Faixa plausível de umidade de grão. O export traz muitas leituras inválidas:
# sensor desligado (0 ou ~0,000001), lixo (<5%) e erros (>35%). Ponderar tudo
# derrubava a média (ex.: CA02 do milho dava 5% em vez de ~19%).
UMID_MIN, UMID_MAX = 5.0, 35.0


def vin_valido(v):
    """VIN real é alfanumérico sem espaço (1CQS790AVN0145227). O export usa o próprio
    texto ('Máquina Desconhecida', 'CA S670 CONFG') quando não identifica a máquina."""
    return bool(v) and bool(re.fullmatch(r'[A-Za-z0-9]{10,}', str(v).strip()))


def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    # mapeia índice da coluna -> chave (primeira ocorrência de cada cabeçalho)
    idx = {}
    for j, h in enumerate(header):
        k = COLS.get(norm(h))
        if k and k not in idx.values():
            idx[j] = k
    faltando = set(COLS.values()) - set(idx.values())
    if faltando:
        raise SystemExit('Colunas ausentes na planilha: %s' % ', '.join(sorted(faltando)))

    por_maq = defaultdict(lambda: {
        'area': 0.0, 'tempo': 0.0, 'comb': 0.0,
        'umid_x_area': 0.0, 'area_umid': 0.0,     # umidade só sobre linhas com leitura válida
        'vel_x_tempo': 0.0, 'tempo_vel': 0.0,     # idem p/ velocidade
        'ops': 0, 'apelido': None, 'vin': None,
        'op_area': defaultdict(float), 'talhoes': set(),
    })
    por_operador = defaultdict(lambda: {'area': 0.0, 'talhoes': set(), 'tempo': 0.0, 'comb': 0.0})
    culturas = set()
    num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
    ignorado = defaultdict(float)  # máquinas sem VIN real -> ha (não entram na frota)

    for r in rows:
        if r is None or all(c is None for c in r):
            continue
        d = {k: r[j] for j, k in idx.items() if j < len(r)}
        area, tempo, comb = num(d.get('area')), num(d.get('tempo')), num(d.get('combustivel'))
        if area <= 0:
            continue
        culturas.add(str(d.get('cultura') or '').strip())
        op = str(d.get('operador') or '').strip()

        # operador conta SEMPRE (o trabalho existiu, mesmo sem máquina identificada)
        if op and op != '---':
            o = por_operador[op]
            o['area'] += area
            o['tempo'] += tempo
            o['comb'] += comb
            if d.get('talhao'):
                o['talhoes'].add(str(d['talhao']).strip())

        vin = d.get('vin')
        if not vin_valido(vin):           # 'Máquina Desconhecida', 'CA S670 CONFG'...
            ignorado[str(d.get('apelido') or vin or '?').strip()] += area
            continue

        m = por_maq[str(vin).strip()]
        m['area'] += area
        m['tempo'] += tempo
        m['comb'] += comb
        u = num(d.get('umidade'))
        if UMID_MIN <= u <= UMID_MAX:      # só leitura plausível (ver UMID_MIN/UMID_MAX)
            m['umid_x_area'] += u * area
            m['area_umid'] += area
        v = num(d.get('velocidade'))
        if v > 0 and tempo > 0:
            m['vel_x_tempo'] += v * tempo
            m['tempo_vel'] += tempo
        m['ops'] += 1
        m['apelido'] = m['apelido'] or d.get('apelido')
        m['vin'] = str(vin).strip()
        if op and op != '---':
            m['op_area'][op] += area
        if d.get('talhao'):
            m['talhoes'].add(str(d['talhao']).strip())

    wb.close()

    maquinas = []
    for vin, m in por_maq.items():
        hrs = m['tempo'] / 3600.0
        dominante = max(m['op_area'].items(), key=lambda x: x[1])[0] if m['op_area'] else None
        maquinas.append({
            'maq': ca_from_apelido(m['apelido']),
            'apelido': m['apelido'],
            'vin': vin,
            'operador': dominante,
            'operadores': sorted(m['op_area'], key=m['op_area'].get, reverse=True),
            'ops': m['ops'],
            'talhoes': len(m['talhoes']),
            'ha': round(m['area'], 1),
            'haHr': round(m['area'] / hrs, 1) if hrs > 0 else None,          # rendimento
            'umidade': round(m['umid_x_area'] / m['area_umid'], 1) if m['area_umid'] else None,
            # % da área da máquina que tem leitura de umidade válida (transparência)
            'umidCobertura': round(100 * m['area_umid'] / m['area']) if m['area'] else 0,
            'lha': round(m['comb'] / m['area'], 1) if m['area'] else None,   # consumo por área
            'lh': round(m['comb'] / hrs, 1) if hrs > 0 else None,            # consumo por hora
            'velocidade': round(m['vel_x_tempo'] / m['tempo_vel'], 1) if m['tempo_vel'] else None,
            'horas': round(hrs, 1),
            'combustivel': round(m['comb'], 1),
        })
    maquinas.sort(key=lambda x: -(x['ha'] or 0))

    operadores = []
    for op, o in por_operador.items():
        hrs = o['tempo'] / 3600.0
        operadores.append({
            'operador': op,
            'ha': round(o['area'], 1),
            'talhoes': len(o['talhoes']),
            'haHr': round(o['area'] / hrs, 1) if hrs > 0 else None,
            'lh': round(o['comb'] / hrs, 1) if hrs > 0 else None,
        })
    operadores.sort(key=lambda x: -x['ha'])

    return {
        'culturas': sorted(c for c in culturas if c),
        'maquinas': maquinas,
        'operadores': operadores,
        'totalHa': round(sum(m['ha'] for m in maquinas), 1),
        # área que existe no export mas ficou fora da frota (máquina não identificada)
        'semMaquina': {k: round(v, 1) for k, v in sorted(ignorado.items(), key=lambda x: -x[1])},
    }


if __name__ == '__main__':
    if len(sys.argv) < 3:
        raise SystemExit('uso: parse_opc_frota.py <entrada.xlsx> <saida.json>')
    out = parse(sys.argv[1])
    with open(sys.argv[2], 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print('[parse_opc_frota] %s -> %s' % (sys.argv[1], sys.argv[2]))
    print('  culturas: %s' % ', '.join(out['culturas']))
    print('  maquinas: %d | operadores: %d | area total: %.1f ha'
          % (len(out['maquinas']), len(out['operadores']), out['totalHa']))
    if out['semMaquina']:
        print('  FORA da frota (sem VIN): %s'
              % ', '.join('%s=%.1f ha' % (k, v) for k, v in out['semMaquina'].items()))
    for m in out['maquinas'][:3]:
        print('   %s ha=%s haHr=%s umid=%s lha=%s lh=%s vel=%s op=%s'
              % (m['maq'], m['ha'], m['haHr'], m['umidade'], m['lha'], m['lh'], m['velocidade'], m['operador']))
